import openpyxl
from openpyxl.styles import PatternFill, NamedStyle
import pandas as pd

# Leitura e tratamento dos dados do CSV
file_path = 'path/to/your/inter_file.csv'
data = pd.read_csv(file_path, skiprows=4, delimiter=';', decimal=',', thousands='.')
data['Valor'] = pd.to_numeric(data['Valor'].apply(lambda x: x.replace('.', '').replace(',', '.') if isinstance(x, str) else x), errors='coerce')
entradas = data[data['Valor'] > 0]
saidas = data[data['Valor'] < 0]

# Inicialização da planilha Excel
wb = openpyxl.Workbook()
sheet = wb.active

# Definição das cores
header_fill_entradas = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
header_fill_saidas = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
header_fill_resumo = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

# Definindo os estilos de data e moeda
date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
currency_style = NamedStyle(name='currency_style', number_format='"R$"#,##0.00')

# Títulos e cabeçalhos
sheet.merge_cells('A1:D1')
sheet['A1'] = 'Entradas'
sheet.merge_cells('F1:I1')
sheet['F1'] = 'Saídas'
sheet.merge_cells('K1:M1')
sheet['K1'] = 'Resumo'
headers = ['Data', 'Nome', 'Tipo', 'Valor']
sheet.append(headers + ['']  + headers + ['']  + ['Somatório Entradas', 'Somatório Saídas', 'Restante'])

# Preenchimento dos dados
for i, row in enumerate(entradas.itertuples(index=False), start=3):
    sheet.cell(row=i, column=1, value=row[0]).style = date_style
    sheet.cell(row=i, column=2, value=row[2])
    sheet.cell(row=i, column=3, value=row[1])
    sheet.cell(row=i, column=4, value=row[3]).style = currency_style

for i, row in enumerate(saidas.itertuples(index=False), start=3):
    sheet.cell(row=i, column=6, value=row[0]).style = date_style
    sheet.cell(row=i, column=7, value=row[2])
    sheet.cell(row=i, column=8, value=row[1])
    sheet.cell(row=i, column=9, value=abs(row[3])).style = currency_style

# Fórmulas de somatório e restante
last_row = sheet.max_row
sheet['K3'] = f'=SUM(D3:D{last_row})'
sheet['L3'] = f'=SUM(I3:I{last_row})'
sheet['M3'] = '=K3-L3'

# Aplicando estilos de moeda às células de somatório e restante
sheet['K3'].style = currency_style
sheet['L3'].style = currency_style
sheet['M3'].style = currency_style

# Aplicando as formatações de cor
for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=4):
    for cell in row:
        cell.fill = header_fill_entradas
for row in sheet.iter_rows(min_row=1, max_row=2, min_col=6, max_col=9):
    for cell in row:
        cell.fill = header_fill_saidas
for row in sheet.iter_rows(min_row=1, max_row=2, min_col=11, max_col=13):
    for cell in row:
        cell.fill = header_fill_resumo

# Ajustando o tamanho das células ao conteúdo máximo
for col in sheet.columns:
    max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
    sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max_length + 3
    
# Salvando a planilha final corrigida
wb.save('planilha_final.xlsx')